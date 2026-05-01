# -*- coding: utf-8 -*-
"""Load hostnames from CSV or Excel files (column header Hostname)."""

import csv
import os

from openpyxl import load_workbook

HOSTNAME_COLUMN = "Hostname"


def load_hostnames_from_file(path: str) -> list[str]:
    """
    Read non-empty hostname values from the Hostname column.

    :param path: Path to a .csv or .xlsx file
    :return: Hostnames in row order (duplicates preserved)
    :raises ValueError: unsupported extension, missing column, or empty list
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _load_csv(path)
    if ext == ".xlsx":
        return _load_xlsx(path)
    raise ValueError(
        f"Unsupported host list format {ext!r}. Use .csv or .xlsx."
    )


def _load_csv(path: str) -> list[str]:
    out: list[str] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError("CSV file is empty or has no header row.")
        if HOSTNAME_COLUMN not in reader.fieldnames:
            raise ValueError(
                f'CSV must contain a "{HOSTNAME_COLUMN}" column header.'
            )
        for row in reader:
            val = row.get(HOSTNAME_COLUMN, "")
            if val is None:
                continue
            s = str(val).strip()
            if s:
                out.append(s)
    if not out:
        raise ValueError(
            f'No non-empty values under "{HOSTNAME_COLUMN}" in {path!r}.'
        )
    return out


def _load_xlsx(path: str) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first is None:
            raise ValueError("Excel file has no header row.")
        headers = [
            str(c).strip() if c is not None else "" for c in first
        ]
        try:
            idx = headers.index(HOSTNAME_COLUMN)
        except ValueError:
            raise ValueError(
                f'Excel sheet must contain a "{HOSTNAME_COLUMN}" column header.'
            )
        out: list[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cells = list(row) if row is not None else []
            if idx >= len(cells):
                continue
            val = cells[idx]
            if val is None:
                continue
            s = str(val).strip()
            if s:
                out.append(s)
    finally:
        wb.close()
    if not out:
        raise ValueError(
            f'No non-empty values under "{HOSTNAME_COLUMN}" in {path!r}.'
        )
    return out
